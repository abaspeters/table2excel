"""Final step: write the tables out.

Design choices worth keeping:

* One table per sheet, so nothing downstream has to guess where one table
  ends and the next begins.
* An _Index sheet carrying provenance — file, page range, image-quality notes.
  Six months from now the first question anyone asks about a number is where
  it came from.
* Every cell the extractor could not read is given a faint fill and a comment.
  The highlight is keyed off the cell's TEXT, not off a position recorded at
  extraction time, so it survives the review step: fix a cell in the editor
  and the highlight disappears on its own; miss one and it stays lit.
"""

import re
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table as XLTable, TableStyleInfo

from .models import UNREADABLE, Table

FONT = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="1F3864")

# Faint amber. Strong enough to catch the eye when scanning a sheet, light
# enough that the text stays readable and printing it doesn't waste toner.
UNREADABLE_FILL = PatternFill("solid", fgColor="FDE9D9")
UNREADABLE_BORDER = Border(*[Side(style="thin", color="E08A2E")] * 4)
UNREADABLE_FONT = Font(name=FONT, color="9C4A00", bold=True)

FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")
NUMERIC = re.compile(r"^-?[\d,]*\.?\d+%?$")


def _safe_sheet_name(name: str, used: set) -> str:
    name = re.sub(r"[\[\]:*?/\\]", "-", name)[:28] or "Table"
    candidate, i = name, 1
    while candidate.lower() in used:
        i += 1
        candidate = f"{name[:25]}_{i}"
    used.add(candidate.lower())
    return candidate


def _coerce(value: str):
    """Turn text that is unambiguously a number into a real number, plus the
    display format that preserves how it looked on the page.

    Anything with a currency symbol, a stray letter or an illegibility marker
    is left as text on purpose — a wrong number is worse than a string.
    """
    v = str(value).strip()
    if not v or not NUMERIC.match(v):
        return value, None
    try:
        if v.endswith("%"):
            body = v[:-1].replace(",", "")
            dp = len(body.split(".")[1]) if "." in body else 0
            return float(body) / 100, f"0.{'0' * dp}%" if dp else "0%"
        body = v.replace(",", "")
        grouped = "," in v
        if "." in body:
            dp = len(body.split(".")[1])
            return float(body), ("#,##0." if grouped else "0.") + "0" * dp
        return int(body), "#,##0" if grouped else None
    except ValueError:
        return value, None


def _mark_unreadable(cell, table: Table) -> None:
    cell.fill = UNREADABLE_FILL
    cell.border = UNREADABLE_BORDER
    cell.font = UNREADABLE_FONT
    cell.comment = Comment(
        f"Not recognised in the source scan.\n"
        f"Check {table.source_file}, page {table.page_range}, and type the "
        f"correct value over this cell.\n"
        f"The extractor was instructed never to guess, so nothing has been "
        f"invented here.",
        "Table extractor", width=280, height=110)


def write(tables: Iterable[Table], out_path: str) -> str:
    tables: List[Table] = list(tables)
    wb = Workbook()
    wb.remove(wb.active)

    index = wb.create_sheet("_Index")
    index.append(["Sheet", "Table title", "Source file", "Page(s)",
                  "Extraction method", "Rows", "Cols", "Unreadable cells",
                  "Notes"])

    used, total_unreadable = set(), 0

    for n, t in enumerate(tables, start=1):
        base = t.title or f"p{t.page_range}_table{n}"
        sheet_name = _safe_sheet_name(base, used)
        ws = wb.create_sheet(sheet_name)

        ws.append(t.headers)
        sheet_unreadable = 0

        for row in t.rows:
            coerced = [_coerce(c) for c in row]
            ws.append([v for v, _ in coerced])
            r = ws.max_row
            for col, ((value, fmt), original) in enumerate(zip(coerced, row), start=1):
                cell = ws.cell(row=r, column=col)
                if UNREADABLE in str(original):
                    _mark_unreadable(cell, t)
                    sheet_unreadable += 1
                    continue
                cell.font = Font(name=FONT)
                if fmt:
                    cell.number_format = fmt

        total_unreadable += sheet_unreadable

        for cell in ws[1]:
            cell.font = Font(name=FONT, bold=True, color="FFFFFF")
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        if t.n_rows and t.n_cols:
            ref = f"A1:{get_column_letter(t.n_cols)}{t.n_rows + 1}"
            xl = XLTable(displayName="tbl_" + re.sub(r"\W", "_", sheet_name), ref=ref)
            xl.tableStyleInfo = TableStyleInfo(name="TableStyleLight9",
                                               showRowStripes=True)
            ws.add_table(xl)

        for i in range(1, t.n_cols + 1):
            longest = max([len(str(t.headers[i - 1]))] +
                          [len(str(r[i - 1])) for r in t.rows] or [10])
            ws.column_dimensions[get_column_letter(i)].width = min(max(longest + 2, 10), 45)
        ws.freeze_panes = "A2"

        index.append([sheet_name, t.title or "", t.source_file, t.page_range,
                      t.method, t.n_rows, t.n_cols, sheet_unreadable or "",
                      "; ".join(t.flags)])
        row = index.max_row
        if sheet_unreadable:
            index.cell(row=row, column=8).fill = UNREADABLE_FILL
            index.cell(row=row, column=8).font = UNREADABLE_FONT
        if t.flags:
            index.cell(row=row, column=9).fill = FLAG_FILL

    for cell in index[1]:
        cell.font = Font(name=FONT, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
    for row in index.iter_rows(min_row=2):
        for cell in row:
            if cell.font.name != FONT:
                cell.font = Font(name=FONT)
    for i, w in enumerate([22, 26, 28, 10, 16, 8, 8, 16, 50], start=1):
        index.column_dimensions[get_column_letter(i)].width = w
    index.freeze_panes = "A2"

    _legend(index, len(tables), total_unreadable)
    wb.save(out_path)
    return out_path


def _legend(index, n_tables: int, total_unreadable: int) -> None:
    """A key, two rows below the index. Someone opening this file next year
    needs to know what the shading means without asking anybody."""
    row = index.max_row + 3

    index.cell(row=row, column=1, value="How to read this workbook").font = Font(
        name=FONT, bold=True, size=12)

    sample = index.cell(row=row + 2, column=1, value=UNREADABLE)
    sample.fill = UNREADABLE_FILL
    sample.border = UNREADABLE_BORDER
    sample.font = UNREADABLE_FONT
    index.cell(row=row + 2, column=2, value=(
        "The extractor could not read this cell from the scan. It was told "
        "never to guess, so the value is missing rather than wrong — check "
        "the source page and type it in. Hover any shaded cell for its page "
        "reference.")).font = Font(name=FONT)

    lines = [
        ("Unshaded cells", "Read from the scan with confidence. Still worth "
                           "spot-checking against the source — no extractor is "
                           "perfect on photographed pages."),
        ("Numbers vs text", "Values that were unambiguously numeric are stored "
                            "as real numbers and will sum correctly. Anything "
                            "carrying a currency symbol or a stray character was "
                            "deliberately left as text rather than guessed at."),
        ("Notes column", "Image-quality findings for each page — resolution, "
                         "lighting correction, tilt correction. A low-resolution "
                         "note means treat that sheet's digits with extra care."),
    ]
    for i, (label, text) in enumerate(lines, start=1):
        index.cell(row=row + 2 + i, column=1, value=label).font = Font(name=FONT, bold=True)
        index.cell(row=row + 2 + i, column=2, value=text).font = Font(name=FONT)

    summary = (f"{n_tables} table(s) extracted. "
               f"{total_unreadable} cell(s) need a human."
               if total_unreadable else
               f"{n_tables} table(s) extracted. No unreadable cells.")
    cell = index.cell(row=row + 6 + len(lines), column=1, value=summary)
    cell.font = Font(name=FONT, bold=True,
                     color="9C4A00" if total_unreadable else "1F3864")
